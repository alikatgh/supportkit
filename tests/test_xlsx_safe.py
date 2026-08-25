"""Guards for the malformed-stylesheet loader.

The bug being guarded is not hypothetical: the real source workbook cannot be
opened by openpyxl at all, and the whole project would have stopped there. The
fixture below reproduces the exact defect (an empty ``<fill/>`` in
``xl/styles.xml``) rather than mocking it, so if openpyxl ever starts tolerating
it, ``test_openpyxl_still_rejects_the_defect`` goes red and tells us the
workaround can be retired.
"""

from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from supportkit.xlsx_safe import repaired_copy, safe_workbook_path, styles_need_repair


def _clean_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(path)
    return path


def _break_styles(src: Path, dest: Path) -> Path:
    """Copy ``src`` with one extra, empty ``<fill/>`` — the real file's defect."""
    with zipfile.ZipFile(src) as z:
        members = {n: z.read(n) for n in z.namelist()}
    styles = members["xl/styles.xml"].decode("utf-8")
    start = styles.find("<fills")
    count_end = styles.find(">", start)
    head = styles[start:count_end]
    n = int(head.split('count="')[1].split('"')[0])
    styles = styles.replace(head, head.replace(f'count="{n}"', f'count="{n + 1}"'), 1)
    styles = styles.replace("</fills>", "<fill/></fills>", 1)
    members["xl/styles.xml"] = styles.encode("utf-8")
    with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in members.items():
            out.writestr(name, data)
    return dest


@pytest.fixture
def broken(tmp_path: Path) -> Path:
    return _break_styles(_clean_workbook(tmp_path / "clean.xlsx"), tmp_path / "broken.xlsx")


def test_openpyxl_still_rejects_the_defect(broken: Path) -> None:
    """The premise of this whole module. If this passes, the workaround is dead code."""
    with pytest.raises(TypeError):
        load_workbook(broken)


def test_defect_is_detected(broken: Path, tmp_path: Path) -> None:
    assert styles_need_repair(broken) is True
    assert styles_need_repair(_clean_workbook(tmp_path / "ok.xlsx")) is False


def test_repaired_copy_opens_and_keeps_the_values(broken: Path, tmp_path: Path) -> None:
    dest, n = repaired_copy(broken, tmp_path / "fixed.xlsx")
    assert n == 1
    ws = load_workbook(dest, data_only=True).active
    assert [list(r) for r in ws.iter_rows(values_only=True)] == [["a", "b"], [1, 2]]


def test_source_file_is_never_modified(broken: Path, tmp_path: Path) -> None:
    """`data/` is source data. A loader that edits it in place would be a disaster."""
    before = broken.read_bytes()
    safe_workbook_path(broken, tmp_path / "work")
    assert broken.read_bytes() == before


def test_clean_file_is_passed_through_untouched(tmp_path: Path) -> None:
    clean = _clean_workbook(tmp_path / "ok.xlsx")
    assert safe_workbook_path(clean, tmp_path / "work") == clean


def test_unrelated_corruption_is_not_swallowed(tmp_path: Path) -> None:
    """Only the empty-`<fill/>` case is handled; anything else must still raise."""
    clean = _clean_workbook(tmp_path / "ok.xlsx")
    junk = tmp_path / "junk.xlsx"
    shutil.copyfile(clean, junk)
    junk.write_bytes(junk.read_bytes()[:200])
    with pytest.raises((zipfile.BadZipFile, KeyError, ValueError)):
        load_workbook(safe_workbook_path(junk, tmp_path / "work"))
