"""``supportkit profile <file>`` — drop a sheet, learn what it can answer.

The whole phase-1 promise in one command: read an unknown xlsx/csv, profile
every column, run the date-parser tournament, propose a role mapping with the
measured evidence for each choice, list the defects a human must rule on, and
print the plan — what would run, and what is refused with the reason.

It changes nothing and spends nothing. The mapping it proposes is written only
when asked (``--mapping-out``), and confirming it is a human step by design.
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import UTC, datetime
from pathlib import Path

from .contract import propose_mapping
from .dateparse import get_parser
from .profiling import blank, column_values, month_coverage, profile_sheet
from .registry import plan
from .xlsx_safe import safe_workbook_path


def read_table(path: Path, sheet: str | None, work_dir: Path) -> tuple[str, list[tuple]]:
    """Rows from an xlsx sheet or a csv. For xlsx with no sheet named, the
    sheet with the most rows wins — and the choice is reported, not silent."""
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as fh:
            return path.stem, [tuple(row) for row in csv.reader(fh)]
    from openpyxl import load_workbook
    workbook = load_workbook(safe_workbook_path(path, work_dir), read_only=True, data_only=True)
    if sheet is not None:
        if sheet not in workbook.sheetnames:
            raise SystemExit(f"[error] no sheet named {sheet!r}; sheets: {', '.join(workbook.sheetnames)}")
        ws = workbook[sheet]
        return sheet, list(ws.iter_rows(values_only=True))
    best_name, best_rows = "", []
    for ws in workbook.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > len(best_rows):
            best_name, best_rows = ws.title, rows
    return best_name, best_rows


def gather_facts(mapping, profile, body) -> dict:
    """The measured facts the registry's predicates read. Computed, never
    assumed — a predicate on a guess would gate analyses on fiction."""
    facts: dict[str, object] = {}
    date_col = mapping.role("date")
    if date_col and mapping.date_parser:
        col = profile.column(date_col)
        parser = get_parser(mapping.date_parser)
        pairs = []
        for value in column_values(body, col.index):
            if blank(value):
                continue
            readings = parser.parse(value)
            if readings:
                pairs.append((readings[0].month, readings[0].day))
        coverage = month_coverage(pairs, reference_year=datetime.now(UTC).year)
        facts["complete_months"] = sum(1 for c in coverage if not c.partial)
        facts["month_coverage"] = coverage
    uid_col = mapping.role("user_id")
    if uid_col:
        col = profile.column(uid_col)
        values = [str(v).strip() for v in column_values(body, col.index) if not blank(v)]
        numeric = sum(1 for v in values if v.replace(" ", "").replace("\n", "").isdigit())
        facts["user_id_numeric_share"] = numeric / len(values) if values else 0.0
    return facts


def _print_report(source: Path, sheet: str, profile, proposal, facts, the_plan) -> None:
    w = print
    w(f"supportkit profile — {source.name} · sheet {sheet!r}")
    w(f"{profile.data_rows} data rows below header row {profile.header_row}; "
      f"{profile.dupe_rows} duplicate rows; {len(profile.marker_rows)} instruction rows excluded")
    w("")
    w("COLUMNS")
    for col in profile.columns:
        tops = "; ".join(f"{v[:26]}×{n}" for v, n in col.top[:3]) or "—"
        w(f"  {col.name[:38]:<40} {col.kind:<12} {100 * col.null_rate:>5.1f}% blank  "
          f"{col.cardinality:>5} distinct  {tops}")
    w("")
    w("PROPOSED MAPPING  (a proposal, not a decision — confirm before relying on it)")
    if proposal.mapping.columns:
        for role, column in proposal.mapping.columns.items():
            w(f"  {role:<10} -> {column}")
            w(f"               {proposal.evidence.get(role, '')}")
    else:
        w("  nothing could be proposed — no column met any role's evidence bar")
    losers = [s for s in proposal.date_scores[1:] if s.parsed]
    if losers:
        w("  date parser tournament, losing candidates: " + "; ".join(s.evidence() for s in losers))
    w("")
    coverage = facts.get("month_coverage") or []
    partial = [c for c in coverage if c.partial]
    if partial:
        w("DEFECTS THAT NEED A RULING")
        for c in partial:
            w(f"  month {c.month:02d} is PARTIAL — days {c.first_day}–{c.last_day} of "
              f"{c.days_in_month}; raw counts for it will mislead")
        w("")
    w("WHAT WOULD RUN")
    for analysis in the_plan.runnable:
        w(f"  + {analysis.title}")
    w("")
    w("WHAT YOUR DATA CANNOT ANSWER (and why)")
    for refusal in the_plan.refused:
        w(f"  - {refusal.sentence()}")
    w("")
    w("Sample values above are pseudonymised. Personal names are not — see the pii docs.")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(prog="supportkit", description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)
    prof = sub.add_parser("profile", help="profile a ticket sheet and print the plan")
    prof.add_argument("source", type=Path)
    prof.add_argument("--sheet", default=None)
    prof.add_argument("--mapping-out", type=Path, default=None,
                      help="write the proposed mapping as JSON for the wizard / a later run")
    prof.add_argument("--work-dir", type=Path, default=Path(".supportkit_work"),
                      help="where a repaired copy of a malformed workbook goes")
    args = parser.parse_args(argv)

    if not args.source.exists():
        print(f"[error] no such file: {args.source}", file=sys.stderr)
        return 2

    sheet, rows = read_table(args.source, args.sheet, args.work_dir)
    profile, body, headers = profile_sheet(sheet, rows)
    values_by_column = {c.name: column_values(body, c.index) for c in profile.columns}
    proposal = propose_mapping(profile, values_by_column)
    facts = gather_facts(proposal.mapping, profile, body)
    the_plan = plan(proposal.mapping, facts)
    _print_report(args.source, sheet, profile, proposal, facts, the_plan)

    if args.mapping_out:
        proposal.mapping.save(args.mapping_out)
        print(f"[ok] proposed mapping written to {args.mapping_out} — edit and confirm it; "
              f"it is a proposal, not a decision")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
